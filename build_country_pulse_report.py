import openpyxl, re, pickle, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SB_PATH='/mnt/user-data/uploads/Survey_Basics__What___Why.xlsx'
OUT='/mnt/user-data/outputs'
COUNTRY='POLAND 2026'
COUNTRY_NAME=re.sub(r'\s*\d{4}\s*','',COUNTRY).strip().title()
COUNTRY_FILE=re.sub(r'[^A-Za-z0-9]+','_',COUNTRY_NAME).strip('_')
os.makedirs(OUT,exist_ok=True)

wb_sb=openpyxl.load_workbook(SB_PATH,data_only=True)

# Load approved corrections
import json
CORRECTIONS_PATH = '/home/claude/Survey_Basics_Corrections.json'
try:
    CORRECTIONS = json.load(open(CORRECTIONS_PATH))['corrections']
except:
    CORRECTIONS = {}

def norm(s):
    if s is None: return ''
    return re.sub(r'[^a-z0-9]+',' ',str(s).lower()).strip()[:80]

def norm60(s):
    """60-char key for corrections lookup"""
    if s is None: return ''
    return re.sub(r'[^a-z0-9]+',' ',str(s).lower()).strip()[:60]
SBROW={}
for sh in ['HR','L&D','L&C','MPD','Counseling','Women','Singles','Marriages','JVK']:
    ws=wb_sb[sh]; d={}
    for r in range(3,ws.max_row+1):
        b=ws.cell(row=r,column=2).value
        if b is not None: d[norm(b)]=r
    SBROW[sh]=d
def sb_whatlearn(sheet,qtext,level):
    # Check manual interpretations first (questions not in Survey Basics)
    key_low = norm(qtext)[:60]
    for fragment, levels in MANUAL_INTERP.items():
        if fragment in key_low:
            return levels.get(level, levels.get('med',''))
    # Check approved corrections first
    corr_key = norm60(qtext) + '__' + level
    if corr_key in CORRECTIONS:
        return CORRECTIONS[corr_key]['correction']
    # Fall back to Survey Basics file
    col={'low':6,'med':9,'high':12}[level]
    ws=wb_sb[sheet]; key=norm(qtext); row=SBROW[sheet].get(key)
    if row is None:
        for k,v in SBROW[sheet].items():
            if k[:40]==key[:40]: row=v; break
    if row is None: return '[Survey Basics text not found]'
    return str(ws.cell(row=row,column=col).value or '').strip()

# ── Hybrid scoring scale classifications ──────────────────────────────────
# 'dist' = structural/process/access questions → three-factor distribution scale
# 'mean' = personal/relational/formation questions → mean scale
# Small n (<=5) always falls back to mean regardless of scale setting
# Questions not in Survey Basics — manual interpretation text by question fragment
MANUAL_INTERP = {
    "our organization provides opportunities for women to encoura": {
        'low':  'Staff report limited or inconsistent opportunities for women to connect and support one another.',
        'med':  'Some opportunities exist for women to encourage one another, but they are not consistently accessible or utilized.',
        'high': 'Staff report strong organizational structures that enable women to encourage and support one another.',
    },
}

SCALE_MAP = {
    # SINGLES
    "I have access to resources that address the unique needs":                     'dist',
    "I have a clear understanding of what is expected of me in ministry":           'dist',
    "My practical needs as a single (housing, financial, social)":                 'dist',
    "I feel my singleness is respected and valued":                                'mean',
    "I am learning to navigate my singleness":                                     'mean',
    "I feel relationally connected to my team and community as a single":          'mean',
    "I see my gifts and opportunities as a single person":                         'mean',
    "I have safe people I can turn to for support":                                'mean',
    "I sometimes feel the weight of carrying ministry responsibilities":            'mean',
    "I sometimes feel emotionally or spiritually depleted":                        'mean',
    # HR
    "I believe that HR policies and decisions are applied fairly":                 'dist',
    "HR processes and systems are clear and efficient":                            'dist',
    "I often feel unsure about my place within the organization":                  'dist',
    "I often feel confused or overwhelmed by complicated HR requirements":         'dist',
    "The compensation and benefits I receive are appropriate":                     'dist',
    "I am able to utilize HR information, tools, and the support":                 'dist',
    "I have a clear and up-to-date Position Focus":                               'dist',
    "I have a working knowledge of JV policies and procedures":                   'mean',  # knowledge/confidence
    "I feel noticed and cared for by my team when I have needs":                  'mean',
    # JVK
    "My children are growing in resilience":                                       'mean',
    "JV provides opportunities for my kids to connect":                            'dist',
    "I see my children thriving in at least some areas":                          'mean',
    "I regularly feel my children's needs are overlooked":                        'mean',
    "My children often feel isolated or disconnected":                             'mean',
    "I am aware of available resources to support my children":                   'dist',
    "I have someone to turn to for help when my kids":                            'dist',
    "I clearly understand JV's approach to caring for kids":                      'dist',
    "I feel my children are cared for and supported by JV":                       'mean',
    "My children have 1-2 adults outside our family":                             'mean',
    # COUNSELING
    "I understand JV's process for getting counseling help":                      'dist',
    "I feel encouraged to pursue counseling when needed":                         'mean',
    "I feel more equipped to navigate challenges because of counseling":          'mean',
    "I know who to contact for personal or family care":                          'dist',
    "Counseling is viewed in our organization as a healthy and constructive":     'mean',
    "I have safe and trusted people I can talk to about seeking help":            'mean',
    "I see counseling as a proactive tool for growth":                            'mean',
    "Practical barriers (time, cost, access) keep me from seeking counseling":    'dist',
    "I know someone on staff who has benefitted from counseling":                 'dist',
    # MPD
    "Financial pressure sometimes distracts me from focusing on ministry":        'dist',
    "I often feel alone in carrying the responsibility of MPD":                   'mean',
    "I know who to turn to for encouragement or accountability in my MPD":        'dist',
    "I have the practical MPD tools and guidance I need":                         'dist',
    "I receive valid and regular financial reports about my support team":         'dist',
    "I feel supported by my uplink or ministry team in building":                 'mean',
    "I am confident when sharing my ministry vision and financial needs":          'mean',
    "I am effective when sharing my ministry vision and financial needs":          'mean',
    "I regularly communicate with my partners to let them know":                  'mean',
    # L&D
    "I often struggle to apply Christ's strategy to daily ministry":              'mean',
    "I frequently feel unsure about how to move forward in my development":       'dist',
    "The equipping resources available enable me to be developed":                'dist',
    "I am experiencing professional growth in this season":                       'mean',
    "I receive helpful feedback and encouragement that supports my learning":     'mean',
    "I am growing in healthy rhythms that help me serve others":                  'mean',
    "My uplink rhythms (meetings, guidance, support) help me thrive":             'dist',
    "I am continually learning how Christ's strategy shapes":                     'mean',
    "I am experiencing personal growth in this season":                           'mean',
    # MARRIAGES
    "I feel supported and encouraged by JV and my team culture to prioritize":    'mean',
    "I often feel that ministry demands leave me too depleted to invest in my marriage": 'mean',
    "I know where to go for help if our marriage faces challenges":               'dist',
    "We are learning to navigate ministry pressures together":                    'mean',
    "I have couples or mentors I can turn to for support":                        'mean',
    "My team culture values and respects the importance of nurturing marriages":  'mean',
    # JV WOMEN
    "I struggle to understand how my gifts and role fit":                         'mean',
    "I often feel isolated in ministry and lack women":                           'mean',
    "I often feel disconnected from my team and uninformed":                      'dist',
    "I feel my voice is valued in team and organizational settings":              'mean',
    "I feel that my organization provides clear guidance on women's roles":       'dist',
    "Our organization provides opportunities for women to encourage":             'dist',
    "JV gatherings (such as conferences or retreats) provide a safe":            'mean',
    # L&C
    "I receive regular accountability and helpful feedback on my progress in language": 'dist',
    "I regularly feel discouraged about my pace of language learning":            'mean',
    "I often struggle to balance ministry demands with language and culture":     'mean',
    "I feel increasingly capable in ministry because of my language":             'mean',
    "I know who to turn to for help with language learning challenges":           'dist',
    "I clearly understand the expectations for my progress in language":         'dist',
    "My team helps me with my language and cultural adaptation needs":            'mean',
    "I am growing in my ability to live and function daily in another culture":  'mean',
    "I am aware of cultural differences on my team":                             'mean',
    "It is important for me to grow in my English skills":                       'mean',
    "I am able to switch into English and still communicate effectively":         'mean',
    "I regularly offer grace and encouragement to teammates who experience":      'mean',
}

def get_scale(qtext, n):
    """Look up scale classification by matching question text prefix."""
    if n <= 5:
        return 'mean'  # small n always falls back
    for key, scale in SCALE_MAP.items():
        if key.lower() in qtext.lower():
            return scale
    return 'mean'  # default to mean if not found


def level_of(s): return 'high' if s>=3.50 else ('med' if s>=2.50 else 'low')
def status_of(s): return {'high':'Healthy','med':'Watch','low':'Concern'}[level_of(s)]
def final(raw,burden): return round(6-raw,3) if burden else round(raw,3)

def refined_dist_status(counts, n, is_burden):
    """Three-factor distribution scoring for structural/process questions.
    Healthy  = pos >= 75% AND neg <= 15%
    Watch    = pos >= 50% AND neg <= 30% AND negatives don't outnumber positives
    Concern  = pos < 50% OR neg > 30% OR negatives outnumber positives
    Small n rule: if n <= 5, caller must fall back to mean scale.
    """
    sd,d,u,a,sa = counts
    pos = ((sd+d) if is_burden else (a+sa)) / n * 100
    neg = ((a+sa) if is_burden else (sd+d)) / n * 100
    neg_dominates = neg > pos
    if pos >= 75 and neg <= 15:
        return 'Healthy'
    elif pos >= 50 and neg <= 30 and not neg_dominates:
        return 'Watch'
    else:
        return 'Concern'

def hybrid_status(fs, counts, n, is_burden, scale):
    """Apply correct scale based on question classification.
    scale = 'dist' for structural/process questions (if n > 5)
    scale = 'mean' for personal/relational/formation questions (or any n <= 5)
    """
    if scale == 'dist' and n > 5:
        return refined_dist_status(counts, n, is_burden)
    else:
        return status_of(fs)

# ── Palette ────────────────────────────────────────────────────────────────
NAVY='154360'; HOWFILL='D6E4F0'; HOWTXT='1F4E79'; GREY='444444'
AMBER_FILL='FFF3CD'; AMBER_TXT='854F0B'
GREEN_FILL='EAF3DE'; GREEN_TXT='3B6D11'
QFILL='F5F5F5'; QTXT='154360'; BODY='444444'
INPUT='FFFDE7'; HINT='888888'
SEC_GREEN='3B6D11'; SEC_AMBER='854F0B'; SEC_BLUE='1F4E79'
SEC_QUOTE='444444'; SEC_HEAT='1F3A6E'; HEAT_HDR='1F3A6E'
HEAT_BASE=(0x15,0x65,0xC0)
thin=Side(style='thin',color='E0E0E0')
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

# Culture tag colors
CULTURE_COLORS = {
    '2nd Culture': ('2C3E50','D6E4F0'),   # dark navy text, light blue bg
    '1st Culture': ('145A32','EAF3DE'),   # dark green text, light green bg
    '':            ('888888','F5F5F5'),   # grey for untagged
}

def F(color='444444',sz=10,bold=False,italic=False):
    return Font(name='Calibri',size=sz,bold=bold,italic=italic,color='FF'+color)
def Fill(c): return PatternFill('solid',fgColor='FF'+c)
def AL(h=None,v='center',wrap=False,indent=0):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap,indent=indent)

def heat_fill(frac):
    f=min(frac/0.72,1.0)
    r=int(255-(255-HEAT_BASE[0])*f); g=int(255-(255-HEAT_BASE[1])*f); b=int(255-(255-HEAT_BASE[2])*f)
    return '%02X%02X%02X'%(r,g,b), ('042C53' if f>=0.45 else '444444')

def _lines(text,width,fontsize=10):
    if text is None: return 1
    text=str(text); cpl=max(int(width*(10.0/fontsize)*0.92),4)
    total=0
    for seg in text.split("\n"):
        seg=seg if seg else " "; total+=max(1,-(-len(seg)//cpl))
    return max(total,1)

def autoheight(*specs,pad=12,line_px=15,minh=24,stack=True):
    counts=[_lines(t,w,fs) for t,w,fs in specs]
    need=(sum(counts) if stack else max(counts)) if counts else 1
    return max(minh,max(need,1)*line_px+pad)

HOWTO=[
    '📋  SECTION 1 — QUESTION SCORES: Each question shows the score, status, and the exact Survey Basics interpretation. Column G = flag or correct the interpretation if it does not fit.',
    '✅  SECTIONS 2 & 3 — STRENGTHS & GROWTH AREAS: Column F = type Yes to keep the statement. Column G = type a rewrite if you want different wording.',
    '❓  SECTION 4 — LEADERSHIP QUESTIONS: Column F = type Yes next to the 1–2 questions (maximum) you want carried into the final report.',
    '💬  SECTION 5 — STAFF VOICE QUOTES: Column F = type Yes for the quotes you want included. Quotes appear exactly as written and must not be changed.',
    '🔥  SECTION 6 — RESPONSE HEATMAP: Full response distribution for reference. No input needed.',
]

# ── quote tuple format: (tag, quote_text, culture_label)
# culture_label = '1st Culture' | '2nd Culture' | '' (blank = not tagged)

def build(dept,rank,total,target_ws=None,save=True):
    name=dept['name']; avg=dept['avg']
    # Apply concern-count rule to determine effective department status
    q_statuses_all = []
    for gname, qs in dept['groups']:
        for q in qs:
            sheet,lookup,disp,raw,burden,counts,n = q
            fs = final(raw, burden)
            st = hybrid_status(fs, counts, n, burden, get_scale(disp, n))
            q_statuses_all.append(st)
    concern_count_dept = q_statuses_all.count('Concern')
    status = 'Concern' if concern_count_dept >= 3 else dept['status']
    ncount=max(q[6] for g,qs in dept['groups'] for q in qs)
    has_culture_tags = any(len(item)==3 and item[2] for item in dept['quotes'])

    if target_ws is None:
        wb=openpyxl.Workbook(); ws=wb.active
    else:
        ws=target_ws; wb=None
    ws.title=f'{COUNTRY_NAME} {name}'[:31]
    ws.sheet_view.showGridLines=False
    ws.column_dimensions['A'].width=14.71
    ws.column_dimensions['B'].width=44.71
    ws.column_dimensions['C'].width=11.71
    ws.column_dimensions['D'].width=10
    ws.column_dimensions['E'].width=46
    ws.column_dimensions['F'].width=12
    ws.column_dimensions['G'].width=34
    ws.column_dimensions['H'].width=44
    ws.column_dimensions['I'].width=2
    ws.column_dimensions['J'].width=10
    ws.column_dimensions['K'].width=10
    ws.column_dimensions['L'].width=10
    ws.column_dimensions['M'].width=10
    ws.column_dimensions['N'].width=10
    ws.column_dimensions['O'].width=10
    ws.column_dimensions['P'].width=10

    def merge_bar(r,text,fill,txtcolor,sz=12,bold=True,h=26,italic=False,halign=None):
        ws.merge_cells(f'A{r}:P{r}')
        c=ws.cell(row=r,column=1,value=text)
        c.font=F(txtcolor,sz,bold,italic); c.fill=Fill(fill)
        c.alignment=AL(h=halign,v='center',wrap=False,indent=1 if not halign else 0)
        ws.row_dimensions[r].height=h
        for cc in range(2,17): ws.cell(row=r,column=cc).fill=Fill(fill)

    r=1
    merge_bar(r,f'PEOPLE AND CULTURE DIRECTORS REVIEW — {COUNTRY}   |   {name.upper()}',NAVY,'FFFFFF',16,True,36,halign='center'); r+=1
    merge_bar(r,f'Department Score: {avg:.2f}   |   Status: {status}   |   Respondents: n = {ncount}   |   Rank {rank} of {total} (worst → best)',NAVY,'FFFFFF',11,False,22,halign='center'); r+=1
    ws.freeze_panes='C3'  # freeze top 2 rows + cols A-B so question text stays visible when scrolling right

    # ── SECTION 1 ──────────────────────────────────────────────────────────
    ws.row_dimensions[r].height=6; r+=1
    merge_bar(r,'SECTION 1 — QUESTION SCORES',NAVY,'FFFFFF',12,True,26); r+=1
    merge_bar(r,'Read each question, score, and interpretation. If the interpretation is incorrect, type the full replacement text in column G.',NAVY,'DDEEEE',10,False,22,italic=True); r+=1
    for i,t in enumerate(['Section','Full Question Text','Score','Status',
        'Survey Basics — What This Score Means','','✏ If incorrect, type the full replacement text here',
        '✏ Does the score reflect what you see in the heatmap? If not, note your reasons here.',
        '','','Score','Strongly\nDisagree','Disagree','Unsure','Agree','Strongly\nAgree']):
        c=ws.cell(row=r,column=i+1,value=t); c.font=F('FFFFFF',9,True)
        if i == 8: c.fill=Fill('EEEEEE'); c.font=F(GREY,9,False)  # spacer
        elif i >= 9: c.fill=Fill(HEAT_HDR)  # heatmap side
        else: c.fill=Fill(GREY)  # left side
    ws.row_dimensions[r].height=30; r+=1
    multi=len(dept['groups'])>1
    for gname,qs in dept['groups']:
        if multi:
            merge_bar(r,f'▸ {gname}',QFILL,QTXT,10,True,18); r+=1
        for q in sorted(qs, key=lambda x: (
                {'Concern':0,'Watch':1,'Healthy':2}.get(
                    hybrid_status(final(x[3],x[4]),x[5],x[6],x[4],get_scale(x[2],x[6])),1),
                final(x[3],x[4])
            )):
            sheet,lookup,disp,raw,burden,counts,n=q
            fs=final(raw,burden); st=hybrid_status(fs,counts,n,burden,get_scale(disp,n))
            # Use effective status level for Survey Basics lookup
            lvl={'Healthy':'high','Watch':'med','Concern':'low'}.get(st,'med')
            interp=sb_whatlearn(sheet,lookup,lvl)
            a=ws.cell(row=r,column=1,value=('Burden\n[inv.]' if burden else 'Q'))
            a.font=F(AMBER_TXT if burden else QTXT,8); a.fill=Fill(AMBER_FILL if burden else QFILL)
            a.alignment=AL('center','center',True); a.border=BORDER
            b=ws.cell(row=r,column=2,value=disp); b.font=F(BODY,9); b.fill=Fill('FFFFFF')
            b.alignment=AL(None,'top',True); b.border=BORDER
            stfill={'Healthy':GREEN_FILL,'Watch':AMBER_FILL,'Concern':'F8D7DA'}.get(st,QFILL)
            stcolor={'Healthy':GREEN_TXT,'Watch':AMBER_TXT,'Concern':'9C0006'}.get(st,BODY)
            cc=ws.cell(row=r,column=3,value=round(fs,2)); cc.font=F(stcolor,11,True)
            cc.fill=Fill(stfill); cc.alignment=AL('center','center'); cc.border=BORDER; cc.number_format='0.00'
            dd=ws.cell(row=r,column=4,value=st); dd.font=F(stcolor,9,True)
            dd.fill=Fill(stfill); dd.alignment=AL('center','center'); dd.border=BORDER
            ee=ws.cell(row=r,column=5,value=interp); ee.font=F(BODY,8,italic=True)
            ee.fill=Fill(QFILL); ee.alignment=AL(None,'top',True); ee.border=BORDER
            ws.cell(row=r,column=6).fill=Fill(INPUT); ws.cell(row=r,column=6).border=BORDER
            gg=ws.cell(row=r,column=7,value='Type full replacement text here if incorrect...'); gg.font=F(HINT,9,italic=True)
            gg.fill=Fill(INPUT); gg.alignment=AL(None,'top',True); gg.border=BORDER
            hh=ws.cell(row=r,column=8,value='Note here if not, and the reasons why you disagree.'); hh.font=F(HINT,9,italic=True); hh.fill=Fill(INPUT); hh.alignment=AL(None,'top',True); hh.border=BORDER
            # Spacer col I
            ws.cell(row=r,column=9).fill=Fill('EEEEEE'); ws.cell(row=r,column=9).border=BORDER
            # Heatmap cols J-P
            jj=ws.cell(row=r,column=10,value=('Burden\n[inv.]' if burden else 'Q'))
            jj.font=F(AMBER_TXT if burden else BODY,8)
            jj.fill=Fill(AMBER_FILL if burden else 'EEF3FA'); jj.alignment=AL('center','center',True); jj.border=BORDER
            # Score column next to heatmap
            sc=ws.cell(row=r,column=11,value=round(fs,2))
            sc.font=F(stcolor,11,True); sc.fill=Fill(stfill)
            sc.alignment=AL('center','center'); sc.border=BORDER; sc.number_format='0.00'
            for ci,cnt in enumerate(counts):
                frac=(cnt/n) if n else 0
                pct=f'{frac*100:.1f}% ({cnt})'
                hfill,fontc=heat_fill(frac)
                hcell=ws.cell(row=r,column=12+ci,value=pct); hcell.font=F(fontc,9)
                hcell.fill=Fill(hfill); hcell.alignment=AL('center','center',True); hcell.border=BORDER
            ws.row_dimensions[r].height=autoheight((disp,44.71,9),(interp,46,8),minh=40,stack=False); r+=1
    ws.row_dimensions[r].height=6; r+=1


    # ── SECTION 2 ──────────────────────────────────────────────────────────
    merge_bar(r,'SECTION 2 — STRENGTHS',SEC_GREEN,'FFFFFF',12,True,26); r+=1
    merge_bar(r,'Column F = type Yes to keep. Column G = type your rewrite if you want different wording.',SEC_GREEN,'E8F5EF',10,False,22,italic=True); r+=1
    for i,t in enumerate(['Section','Draft Strength Statement','—','—','—','Include? (type Yes)','✏ Rewrite (type here)']):
        c=ws.cell(row=r,column=i+1,value=t); c.font=F('FFFFFF',9,True); c.fill=Fill(GREY)
        c.alignment=AL('center','center',True); c.border=BORDER
    ws.row_dimensions[r].height=20; r+=1
    for s in dept['strengths']:
        a=ws.cell(row=r,column=1,value='Strength'); a.font=F(GREEN_TXT,9,True); a.fill=Fill(GREEN_FILL); a.alignment=AL('center','center'); a.border=BORDER
        ws.merge_cells(f'B{r}:E{r}')
        b=ws.cell(row=r,column=2,value=s); b.font=F(BODY,10); b.fill=Fill(GREEN_FILL); b.alignment=AL(None,'center',True); b.border=BORDER
        for cc in range(3,6): ws.cell(row=r,column=cc).fill=Fill(GREEN_FILL); ws.cell(row=r,column=cc).border=BORDER
        ws.cell(row=r,column=6).fill=Fill(INPUT); ws.cell(row=r,column=6).border=BORDER
        ws.cell(row=r,column=7).fill=Fill(INPUT); ws.cell(row=r,column=7).border=BORDER
        ws.row_dimensions[r].height=autoheight((s,105,10),minh=30); r+=1
    ws.row_dimensions[r].height=6; r+=1

    # ── SECTION 3 ──────────────────────────────────────────────────────────
    merge_bar(r,'SECTION 3 — GROWTH AREAS',SEC_AMBER,'FFFFFF',12,True,26); r+=1
    merge_bar(r,'Column F = type Yes to keep. Column G = type your rewrite if you want different wording.',SEC_AMBER,'FFF8EC',10,False,22,italic=True); r+=1
    for i,t in enumerate(['Section','Draft Growth Area Statement','—','—','—','Include? (type Yes)','✏ Rewrite (type here)']):
        c=ws.cell(row=r,column=i+1,value=t); c.font=F('FFFFFF',9,True); c.fill=Fill(GREY)
        c.alignment=AL('center','center',True); c.border=BORDER
    ws.row_dimensions[r].height=20; r+=1
    for s in dept['growth']:
        a=ws.cell(row=r,column=1,value='Growth Area'); a.font=F(AMBER_TXT,9,True); a.fill=Fill(AMBER_FILL); a.alignment=AL('center','center'); a.border=BORDER
        ws.merge_cells(f'B{r}:E{r}')
        b=ws.cell(row=r,column=2,value=s); b.font=F(BODY,10); b.fill=Fill(AMBER_FILL); b.alignment=AL(None,'center',True); b.border=BORDER
        for cc in range(3,6): ws.cell(row=r,column=cc).fill=Fill(AMBER_FILL); ws.cell(row=r,column=cc).border=BORDER
        ws.cell(row=r,column=6).fill=Fill(INPUT); ws.cell(row=r,column=6).border=BORDER
        ws.cell(row=r,column=7).fill=Fill(INPUT); ws.cell(row=r,column=7).border=BORDER
        ws.row_dimensions[r].height=autoheight((s,105,10),minh=30); r+=1
    ws.row_dimensions[r].height=6; r+=1

    # ── SECTION 4 ──────────────────────────────────────────────────────────
    merge_bar(r,'SECTION 4 — LEADERSHIP QUESTIONS',SEC_BLUE,'FFFFFF',12,True,26); r+=1
    merge_bar(r,'Column F = type Yes next to 2 questions you want in the final report. Use the write-in rows below if none of the above fit.',SEC_BLUE,'EEF4FB',10,False,22,italic=True); r+=1
    for i,t in enumerate(['Section','Full Leadership Question','—','—','—','Include? (Yes — max 2)','—']):
        c=ws.cell(row=r,column=i+1,value=t); c.font=F('FFFFFF',9,True); c.fill=Fill(GREY)
        c.alignment=AL('center','center',True); c.border=BORDER
    ws.row_dimensions[r].height=20; r+=1
    for s in dept['leadership']:
        a=ws.cell(row=r,column=1,value='Leader Q'); a.font=F(SEC_BLUE,9,True); a.fill=Fill(HOWFILL); a.alignment=AL('center','center'); a.border=BORDER
        ws.merge_cells(f'B{r}:E{r}')
        b=ws.cell(row=r,column=2,value=s); b.font=F(BODY,10); b.fill=Fill(HOWFILL); b.alignment=AL(None,'center',True); b.border=BORDER
        for cc in range(3,6): ws.cell(row=r,column=cc).fill=Fill(HOWFILL); ws.cell(row=r,column=cc).border=BORDER
        ws.cell(row=r,column=6).fill=Fill(INPUT); ws.cell(row=r,column=6).border=BORDER
        ws.cell(row=r,column=7).fill=Fill(HOWFILL); ws.cell(row=r,column=7).border=BORDER
        ws.row_dimensions[r].height=autoheight((s,105,10),minh=30); r+=1

    # Write-in rows for custom leadership questions (2 rows)
    for _wi in range(2):
        ws.row_dimensions[r].height=52
        a=ws.cell(row=r,column=1,value='Write-in'); a.font=F(SEC_BLUE,9,True); a.fill=Fill(HOWFILL); a.alignment=AL('center','center'); a.border=BORDER
        ws.merge_cells(f'B{r}:E{r}')
        b=ws.cell(row=r,column=2,value='Add your own leadership question here if needed...')
        b.font=F(HINT,10,italic=True); b.fill=Fill(INPUT); b.alignment=AL(None,'center',True); b.border=BORDER
        for cc in range(3,6): ws.cell(row=r,column=cc).fill=Fill(INPUT); ws.cell(row=r,column=cc).border=BORDER
        ws.cell(row=r,column=6).fill=Fill(INPUT); ws.cell(row=r,column=6).border=BORDER
        ws.cell(row=r,column=7).fill=Fill(HOWFILL); ws.cell(row=r,column=7).border=BORDER
        r+=1
    ws.row_dimensions[r].height=6; r+=1

    # ── SECTION 5 ──────────────────────────────────────────────────────────
    merge_bar(r,'SECTION 5 — STAFF VOICE QUOTES',SEC_QUOTE,'FFFFFF',12,True,26); r+=1
    note = 'Column F = type Yes for every quote you want included. Quotes appear exactly as written — do not change the text.'
    if has_culture_tags:
        note += ' Column D shows 1st or 2nd culture.'
    merge_bar(r,note,SEC_QUOTE,'F5F5F5',10,False,22,italic=True); r+=1

    # Show the open-ended question prompt that generated these quotes
    if dept.get('open_q'):
        merge_bar(r,f'Responding to: "{dept["open_q"]}"','F5F5F5','444444',9,False,18,italic=True); r+=1

    # Column headers — D shows culture tag if this dept has them
    qt_headers = ['Section','Full Quote Text — do not change','—',
                  'Culture' if has_culture_tags else 'Tag',
                  'Tag' if has_culture_tags else '—',
                  'Include? (type Yes)','—']
    for i,t in enumerate(qt_headers):
        c=ws.cell(row=r,column=i+1,value=t); c.font=F('FFFFFF',9,True); c.fill=Fill(GREY)
        c.alignment=AL('center','center',True); c.border=BORDER
    ws.row_dimensions[r].height=20; r+=1

    for item in dept['quotes']:
        tag        = item[0]
        quote      = item[1]
        culture    = item[2] if len(item) >= 3 else ''
        polish_orig = item[3] if len(item) == 4 else None

        tagfill = GREEN_FILL if tag=='Strength' else AMBER_FILL
        tagtxt  = GREEN_TXT  if tag=='Strength' else AMBER_TXT
        cult_txt, cult_bg = CULTURE_COLORS.get(culture, CULTURE_COLORS[''])

        a=ws.cell(row=r,column=1,value='Quote'); a.font=F(BODY,9,True); a.fill=Fill(QFILL); a.alignment=AL('center','center'); a.border=BORDER
        ws.merge_cells(f'B{r}:C{r}')
        if polish_orig:
            cell_text = '"' + quote + '"' + '\n\nTranslation: "' + polish_orig + '"'
        else:
            cell_text = '"' + quote + '"'
        b=ws.cell(row=r,column=2,value=cell_text); b.font=F(BODY,10,italic=True)
        b.fill=Fill('FFFFFF'); b.alignment=AL(None,'top',True); b.border=BORDER
        ws.cell(row=r,column=3).border=BORDER; ws.cell(row=r,column=3).fill=Fill('FFFFFF')

        if has_culture_tags:
            # Col D = culture badge
            d=ws.cell(row=r,column=4,value=culture if culture else '—')
            d.font=F(cult_txt,9,True); d.fill=Fill(cult_bg)
            d.alignment=AL('center','center'); d.border=BORDER
            # Col E = tag badge
            e=ws.cell(row=r,column=5,value=tag)
            e.font=F(tagtxt,9,True); e.fill=Fill(tagfill)
            e.alignment=AL('center','center'); e.border=BORDER
        else:
            # Col D = tag badge, Col E = spacer
            d=ws.cell(row=r,column=4,value=tag)
            d.font=F(tagtxt,9,True); d.fill=Fill(tagfill)
            d.alignment=AL('center','center'); d.border=BORDER
            e=ws.cell(row=r,column=5); e.fill=Fill(QFILL); e.border=BORDER

        ws.cell(row=r,column=6).fill=Fill(INPUT); ws.cell(row=r,column=6).border=BORDER
        g=ws.cell(row=r,column=7,value='Quote text must not be changed')
        g.font=F(HINT,8,italic=True); g.fill=Fill(QFILL); g.alignment=AL('center','center'); g.border=BORDER
        qtext_for_height = f'"{quote}"' + (f'\n\nOriginal: "{polish_orig}"' if polish_orig else '')
        ws.row_dimensions[r].height=autoheight((qtext_for_height,56,10),minh=44); r+=1
    ws.row_dimensions[r].height=6; r+=1

    # ── SECTION 6 ──────────────────────────────────────────────────────────
    if save and wb is not None:
        safe=re.sub(r'[^A-Za-z0-9]+','_',name).strip('_')
        path=os.path.join(OUT,f'{COUNTRY_FILE}_{rank:02d}_{safe}_Director_Review.xlsx')
        wb.save(path); return os.path.basename(path)
    return None

def build_overview(ws):
    ws.title=f'{COUNTRY_NAME} Summary'[:31]
    ws.sheet_view.showGridLines=False
    for col,w in zip('ABCDEF',[6,28,10,12,8,66]):
        ws.column_dimensions[col].width=w
    from openpyxl.utils import column_index_from_string
    def bar(r,text,fill,txt,sz=12,bold=True,h=26,halign=None,italic=False,span='A:F'):
        a,b=span.split(':')
        ws.merge_cells(f'{a}{r}:{b}{r}')
        c=ws.cell(row=r,column=1,value=text); c.font=F(txt,sz,bold,italic); c.fill=Fill(fill)
        c.alignment=AL(h=halign,v='center',wrap=False,indent=1 if not halign else 0)
        ws.row_dimensions[r].height=h
        for cc in range(1,column_index_from_string(b)+1): ws.cell(row=r,column=cc).fill=Fill(fill)
    r=1
    bar(r,f'PEOPLE AND CULTURE DIRECTORS REVIEW — {COUNTRY}   |   ALL DEPARTMENTS',NAVY,'FFFFFF',16,True,36,halign='center'); r+=1
    bar(r,'Department health overview — sorted worst → best. Use the tabs below for the full department sheets.',NAVY,'FFFFFF',11,False,22,halign='center'); r+=1
    ws.row_dimensions[r].height=6; r+=1
    for i,t in enumerate(['Rank','Department','Score','Status','n','Area to explore further']):
        c=ws.cell(row=r,column=i+1,value=t); c.font=F('FFFFFF',9,True); c.fill=Fill(GREY)
        c.alignment=AL('center','center',True); c.border=BORDER
    ws.row_dimensions[r].height=20; r+=1
    for rank,d in enumerate(DEPTS,1):
        nc=max(q[6] for g,qs in d['groups'] for q in qs)
        # Apply concern-count rule to department status
        q_statuses = []
        for gname, qs in d['groups']:
            for q in qs:
                sheet,lookup,disp,raw,burden,counts,n = q
                fs = final(raw, burden)
                st = hybrid_status(fs, counts, n, burden, get_scale(disp, n))
                q_statuses.append(st)
        concern_count = q_statuses.count('Concern')
        avg_status = d['status']
        if concern_count >= 3:
            effective_status = 'Concern'
        else:
            effective_status = avg_status
        st=effective_status
        stfill={'Healthy':GREEN_FILL,'Watch':AMBER_FILL,'Concern':'F8D7DA'}.get(st,QFILL)
        sttxt={'Healthy':GREEN_TXT,'Watch':AMBER_TXT,'Concern':'9C0006'}.get(st,BODY)
        lead=d.get('lead_question', d['growth'][0] if d['growth'] else '')
        for col,(val,ha) in enumerate([(rank,'center'),(d['name'],None),(round(d['avg'],2),'center'),(effective_status,'center'),(nc,'center'),(lead,None)],1):
            c=ws.cell(row=r,column=col,value=val)
            if col in(3,4): c.font=F(sttxt,11 if col==3 else 9,True); c.fill=Fill(stfill)
            else: c.font=F(BODY,10,bold=(col==2)); c.fill=Fill('FFFFFF')
            c.alignment=AL(ha,'center' if col!=6 else 'top',wrap=(col==6)); c.border=BORDER
            if col==3: c.number_format='0.00'
        ws.row_dimensions[r].height=autoheight((lead,66,10),minh=28,stack=False); r+=1
    ws.row_dimensions[r].height=6; r+=1
    bar(r,'Status: Healthy >= 3.50  •  Watch 2.50–3.49  •  Concern < 2.50.  Burden questions are inverted before scoring.','FFFFFF',HINT,8,False,18,italic=True)
    for cc in range(1,7): ws.cell(row=r,column=cc).fill=PatternFill(fill_type=None)

# ── DEPARTMENT DATA ────────────────────────────────────────────────────────
# Quote tuples: (tag, text) for standard depts
#               (tag, text, culture) for L&C and JVK where culture is known
# Culture values: '1st Culture' | '2nd Culture'

DEPTS=[

  {"name":"Singles","avg":2.99,"status":"Watch",
   "groups":[("Singles",[
     ("Singles","I have access to resources that address the unique needs of single missionaries.","I have access to resources that address the unique needs of single missionaries.",2.00,False,[3,3,3,0,0],9),
     ("Singles","I have a clear understanding of what is expected of me in ministry, team, and community as a single staff member.","I have a clear understanding of what is expected of me in ministry, team, and community as a single staff member.",2.56,False,[0,5,3,1,0],9),
     ("Singles","My practical needs as a single (housing, financial, social) are adequately acknowledged and supported in my context.","My practical needs as a single (housing, financial, social) are adequately acknowledged and supported in my context.",3.22,False,[0,1,5,3,0],9),
     ("Singles","I feel my singleness is respected and valued by JV leadership.","I feel my singleness is respected and valued by JV leadership.",3.33,False,[0,1,4,4,0],9),
     ("Singles","I'm learning to navigate my singleness with growing peace and purpose.","I am learning to navigate my singleness with growing peace and purpose.",3.78,False,[0,0,3,5,1],9),
     ("Singles","I feel relationally connected to my team and community as a single person.","I feel relationally connected to my team and community as a single person.",4.00,False,[0,0,2,5,2],9),
     ("Singles","I see my gifts and opportunities as a single person being used effectively in ministry.","I see my gifts and opportunities as a single person being used effectively in ministry.",4.11,False,[0,1,1,3,4],9),
     ("Singles","I have safe people I can turn to for support, encouragement, and prayer.","I have safe people I can turn to for support, encouragement, and prayer.",4.33,False,[0,0,1,4,4],9),
     ("Singles","I sometimes feel the weight of carrying ministry responsibilities on my own.","I sometimes feel the weight of carrying ministry responsibilities on my own. [Burden — inverted]",4.67,True,[0,0,1,1,7],9),
     ("Singles","I sometimes feel emotionally or spiritually depleted when I am carrying life and ministry responsibilities on my own.","I sometimes feel emotionally or spiritually depleted when I am carrying life and ministry responsibilities on my own. [Burden — inverted]",4.78,True,[0,0,1,0,8],9),
   ])],
   "strengths":[
     'Relational connection is experienced as a genuine source of support for single staff.',
     'Safe and trusted people are accessible within the team and community.',
     'Single staff report growing in peace and purpose in their singleness.',
     'Gifts and ministry opportunities are seen as meaningful and well-placed.',
   ],
   "growth":[
     'The load single staff experience in ministry — and whether structures exist to share it — warrants direct exploration.',
     'Access to resources designed specifically for single missionaries is not being experienced consistently.',
     'Practical acknowledgment and support for the unique needs of single staff is an area of significant concern.',
     'Clarity around what is expected of single staff in ministry, team, and community life is a recurring concern across the team.',
     'Depletion and the weight of carrying responsibilities alone are among the most significant signals in this survey and warrant sustained attention.',
   ],
   "leadership":["How are you staying connected to what single staff are actually experiencing — not just their ministry output?","How does your team engage with single staff to understand what they carry and whether they need more support?","How do single staff experience belonging and visibility on your team — and how do you know?","What do single staff on your team know about JV resources designed for them — and how do you find that out?"],
      "quotes":[
     ("Growth Area","I don't see anything JV does for singles, not counting the singles retreat I only heard about — the previous one happened before I arrived. I think it would be helpful to have articles on serving as a single. Please do not use the term \'lonely people\' — single life does not have to be lonely.","","Nie widzę nic co JV robi dla singli, nie licząc singles retreat o którym tylko słyszałam, bo poprzednio jak był to jeszcze przed moim przyjazdem. Myślę że byłoby pomocne gdyby były artykuły o służeniu jako singiel. Proszę nie używajcie sformułowania \'samotni ludzie\' - życie jako singiel nie musi być samotne."),
     ("Growth Area","Maybe Selah could be available to visit with close friends (perhaps with a limit per year or something).","",None),
     ("Growth Area","I think check Ins are important as well as the ability to debrief things well. In most contexts we don't have anyone we can turn to, maybe outside of our uplink.","",None),
     ("Growth Area","tak samo traktować nasze wyzwania jak wyzwania małżeństw i rodzin z dziećmi","","Treat our challenges the same as the challenges of married couples and families with children."),
   ],
   "open_q":"What would most strengthen JV's support for singles in your context?",
   "lead_question":"The experience of single staff carrying ministry responsibilities and what structures share the load.",
  },

  {"name":"Human Resources","avg":3.24,"status":"Watch",
   "groups":[("Human Resources",[
     ("HR","I believe that HR policies and decisions are applied fairly across the organization.","I believe that HR policies and decisions are applied fairly across the organization.",2.52,False,[4,5,9,3,0],21),
     ("HR","HR processes and systems are clear and efficient, making it easy to get what I need.","HR processes and systems are clear and efficient, making it easy to get what I need.",2.86,False,[2,4,10,5,0],21),
     ("HR","I often feel unsure about my place within the organization.","I often feel unsure about my place within the organization. [Burden — inverted]",2.95,True,[0,11,4,2,4],21),
     ("HR","I often feel confused or overwhelmed by complicated HR requirements.","I often feel confused or overwhelmed by complicated HR requirements. [Burden — inverted]",2.86,True,[1,10,2,7,1],21),
     ("HR","The compensation and benefits I receive are appropriate for the cost of living and demands of my role.","The compensation and benefits I receive are appropriate for the cost of living and demands of my role.",3.29,False,[2,1,8,9,1],21),
     ("HR","I am able to utilize HR information, tools, and the support I need to do my job effectively.","I am able to utilize HR information, tools, and the support I need to do my job effectively.",3.38,False,[1,2,6,12,0],21),
     ("HR","I have a clear and up-to-date Position Focus with regular opportunities to work within my gifting, experience, and calling.","I have a clear and up-to-date Position Focus with regular opportunities to work within my gifting, experience, and calling.",3.48,False,[1,4,2,12,2],21),
     ("HR","I have a working knowledge of JV policies and procedures.","I have a working knowledge of JV policies and procedures.",3.52,False,[0,2,7,11,1],21),
     ("HR","I feel noticed and cared for by my team when I have needs.","I feel noticed and cared for by my team when I have needs.",3.95,False,[0,3,2,9,7],21),
   ])],
   "strengths":[
     'HR is experienced as a steady and trusted source of support.',
     'Staff report a working knowledge of JV policies and procedures.',
   ],
   "growth":[
     'HR processes and systems are not consistently experienced as clear or easy to navigate.',
     'Awareness of what HR offers and how to access it is not consistently reaching staff.',
     'Compensation is not consistently experienced as appropriate for the cost of living and demands of the role.',
     'The experience of fairness in HR decisions is a significant concern and among the most serious patterns in this department.',
     'Clarity around HR requirements would reduce ongoing frustration and time cost.',
   ],
   "leadership":["Do the staff on your team know who to contact in HR and what HR can do for them?","Which HR processes do staff find most confusing or time-consuming, and what would make them simpler?","How are you addressing the perception that HR policies are not always applied consistently?","What do long-tenured staff on your team need that they are not currently receiving from HR?"],
      "quotes":[
     ("Growth Area","Nie jestem pewna czym zajmuje się nasz dział kadr i co jest dostępne dla mnie żeby mi pomóc, więc to już pokazuje jakiś problem.","","I am not sure what our HR department does and what is available to help me — which already shows a problem."),
     ("Growth Area","Simpler guidelines, all info in one place, one contact person, clarity on how 2nd culture missionaries differ from 1st culture missionaries in terms of HR.","",None),
     ("Growth Area","There\'s a lot of things in a lot of different spots. It would be helpful to have this centralized. Also, I sometimes learn about HR things from asking other missionaries which shows there\'s a communication gap.","",None),
     ("Growth Area","Jasny wykaz obowiązków, większa elastyczność (po zmianach strukturalnych dużo rzeczy jest narzuconych niezgodnych z wcześniejszymi ustaleniami).","","A clear list of duties, greater flexibility — after the structural changes many things were imposed that were not consistent with earlier agreements."),
     ("Growth Area","Większa jasność dotycząca systemów emerytalnych ( w tym większe wsparcie kadr w tym zakresie) oraz podatków (często słyszymy, że potrzebujemy własnego prywatnego doradcy, nawet do wypełniania podstawowych dokumentów wymaganych przez organizację).","","Greater clarity about retirement systems (including more HR support) and taxes. We often hear we need our own private advisor, even for filling out basic documents required by the organization."),
     ("Growth Area","I don\'t know if this is a reasonable or realistic request, but I think it would be really helpful if somehow HR actually knew what I needed and initiated helping me with it — so I can learn how to do it myself.","",None),
     ("Growth Area","Więcej krótkich spotkań lub filmików informacyjnych na temat HR","","More short meetings or informational videos about HR."),
     ("Strength","Right now it\'s hard to say as I really haven\'t had to use HR support that much. I do appreciate the occasional check in :)","",None),
   ],
   "open_q":"What would make HR support more helpful to you?",
   "lead_question":"Clarity and accessibility of HR processes, roles, and support across the team.",
  },

  {"name":"JVK — Josiah Venture Kids","avg":3.29,"status":"Watch",
   "groups":[
     ("Second Culture Parents (n=6)",[
       ("JVK","I'm aware of available resources to support my children in cross-cultural life. (2nd Culture Only)","I am aware of available resources to support my children in cross-cultural life.",2.83,False,[0,2,3,1,0],6),
       ("JVK","I have someone to turn for help when my kids face challenges.","I have someone to turn to for help when my kids face challenges.",3.50,False,[0,1,1,4,0],6),
       ("JVK","I clearly understand JV's approach to caring for kids.","I clearly understand JV's approach to caring for kids.",3.83,False,[0,1,1,2,2],6),
       ("JVK","My children have 1-2 adults outside our family they can talk to if needed.","My children have 1-2 adults outside our family they can talk to if needed.",4.40,False,[0,0,0,3,2],5),
       ("JVK","I feel my children are cared for and supported by JV.","I feel my children are cared for and supported by JV.",4.33,False,[0,0,0,4,2],6),
     ]),
     ("First Culture Parents (n=4)",[
       ("JVK","My children are growing in resilience through our family's ministry context.","My children are growing in resilience through our family's ministry context.",2.00,False,[1,2,1,0,0],4),
       ("JVK","JV provides opportunities for my kids to connect with other kids who share similar experiences.","JV provides opportunities for my kids to connect with other kids who share similar experiences.",2.25,False,[2,0,1,1,0],4),
       ("JVK","I see my children thriving in at least some areas of life.","I see my children thriving in at least some areas of life.",2.25,False,[1,1,2,0,0],4),
       ("JVK","My child/children often feel isolated or disconnected.","My children often feel isolated or disconnected. [Burden — inverted]",3.25,True,[1,0,0,3,0],4),
     ]),
     ("All Parents — Shared Questions (n=10)",[
       ("JVK","I regularly feel my children's needs are overlooked in ministry life.","I regularly feel my children's needs are overlooked in ministry life. [Burden — inverted]",3.90,True,[1,0,2,3,4],10),
       ("JVK","My child/children often feel isolated or disconnected.","My children often feel isolated or disconnected. [Burden — inverted, all parents]",3.20,True,[0,2,3,4,1],10),
       ("JVK","My children are growing in resilience through our family's ministry context.","My children are growing in resilience through our family's ministry context. [all parents]",3.40,False,[1,1,2,4,2],10),
       ("JVK","I'm aware of available resources to support my children in cross-cultural life. (2nd Culture Only)","I am aware of available resources to support my children in cross-cultural life. [all parents]",3.50,False,[0,0,1,6,3],10),
       ("JVK","I feel my children are cared for and supported by JV.","I feel my children are cared for and supported by JV. [all parents]",4.20,False,[2,4,0,2,2],10),
     ]),
   ],
   "strengths":[
     'Second culture families report feeling supported and cared for by JV.',
     'Children of second culture families have trusted adults available outside the home.',
     'JV\'s approach to caring for kids is understood and valued by second culture parents.',
   ],
   "growth":[
     'The experience of first culture families — particularly around resilience, peer connection, and children thriving — reflects patterns that need closer attention.',
     'A significant number of parents report that children\'s needs are at risk of being overlooked in the demands of ministry life.',
     'Peer connection opportunities for first culture kids are not consistently accessible or visible.',
     'Awareness of JVK resources among second culture families is very low — only 17% responded positively.',
     'Clarity around what JVK offers and how to access it would benefit all families.',
   ],
   "leadership":["How are you proactively checking in with first culture families about how their children are doing?","What concrete steps does JVK take to help first culture kids connect with peers who share similar experiences?","How do staff families know what JVK resources are available and how to access them?","What does it mean for a child's needs to be seen and addressed in your ministry context?"],
      "quotes":[
     ("Growth Area","I would love to know what is available. I know about JVK camp and spring conference. But outside of those two things I don\'t know what is formally offered or available for our kids.","2nd Culture",None),
     ("Strength","Mamy taką opiekę w kościele lokalnym, więc nie czuję, że potrzebowalibyśmy tego ze strony JV, chociaż obozy dla dzieci \'narodowych\' byłyby miłą opcją.","1st Culture","We have such care in our local church, so I don\'t feel that we would need this from JV, although camps for nationals\' children would be a nice option."),
     ("Growth Area","My kids are little (1.5 + 3.5), so not all questions were applicable. I think JVK will be helpful in the future, though I wonder if it needs to be a little more geographically limited (like more regional groups).","2nd Culture",None),
   ],
   "open_q":"What would most strengthen JV's care for kids?",
   "lead_question":"Support and connection for first and second culture families — and what each group needs.",
  },

  {"name":"Counseling","avg":3.32,"status":"Watch",
   "groups":[("Counseling",[
     ("Counseling","I understand JV's process for getting counseling help.","I understand JV's process for getting counseling help.",2.05,False,[5,10,4,1,0],20),
     ("Counseling","Practical barriers (time, cost, access) keep me from seeking counseling.","Practical barriers (time, cost, access) keep me from seeking counseling. [Burden — inverted]",3.45,True,[1,7,1,4,7],20),
     ("Counseling","I feel encouraged to pursue counseling when needed.","I feel encouraged to pursue counseling when needed.",3.15,False,[1,5,5,8,1],20),
     ("Counseling","I feel more equipped to navigate challenges because of counseling I have received.","I feel more equipped to navigate challenges because of counseling I have received.",3.40,False,[1,4,4,8,3],20),
     ("Counseling","I know who to contact for personal or family care, especially in times of crisis.","I know who to contact for personal or family care, especially in times of crisis.",3.50,False,[0,5,4,7,4],20),
     ("Counseling","I know someone on staff who has benefitted from counseling.","I know someone on staff who has benefitted from counseling.",3.55,False,[1,2,6,7,4],20),
     ("Counseling","Counseling is viewed in our organization as a healthy and constructive step.","Counseling is viewed in our organization as a healthy and constructive step.",3.70,False,[1,1,5,9,4],20),
     ("Counseling","I have safe and trusted people I can talk to about seeking help.","I have safe and trusted people I can talk to about seeking help.",4.00,False,[1,2,0,10,7],20),
     ("Counseling","I see counseling as a proactive tool for growth, not just crisis.","I see counseling as a proactive tool for growth, not just crisis.",4.00,False,[1,1,3,7,8],20),
   ])],
   "strengths":[
     'Counseling is viewed as a healthy and constructive part of ministry life.',
     'Staff increasingly see counseling as a proactive tool for growth rather than crisis response.',
     'Trusted relationships create a safe environment for seeking help.',
     'Awareness of counseling as a beneficial resource is growing across the team.',
   ],
   "growth":[
     'Clarity around how to access counseling could be significantly strengthened.',
     'Encouragement to pursue counseling could be more consistent across leadership.',
     'Practical barriers to accessing counseling would benefit from more structured support.',
     'The felt impact of counseling could be strengthened through better follow-through pathways.',
   ],
   "leadership":["Could you explain JV's counseling process clearly to a staff member who asks today? If not, that is the starting point.","What practical barriers are keeping staff on your team from accessing counseling?","How consistently are you encouraging staff to pursue counseling — not only in crisis?","What would it take to make the counseling process one step simpler for your team?"],
      "quotes":[
     ("Growth Area","Nie wiem jak działa poradnictwo w JV ani czy i kto jest dostępny w Fali oprócz mojego uplinka. Nie znam jasnej ścieżki szukania pomocy.","","I don\'t know how counseling works in JV or who is available in Fala other than my uplink. I don\'t know the clear pathway for seeking counseling."),
     ("Growth Area","A clear process of how I can sign up for counseling and in what cases.","",None),
     ("Growth Area","To know who in our organization is available to do this and what types of counseling they offer. Or to know a list of recommended organizations outside JV. To know what health insurance covers.","",None),
     ("Growth Area","Regularne przypominanie nam o tym i lista przykładowych tematów, z którymi można się zgłosić do zespołu poradnictwa.","","Regular reminders about this and a list of example topics with which one can approach the counseling team."),
     ("Growth Area","Szkolenie oraz finanse na jego poczet","","Training and funding for it."),
     ("Growth Area","O jaki rodzaj poradnictwa tutaj chodzi? Czy rozmowy mentorskie też w to wchodzą?","","What is meant by counseling here? Do mentoring conversations also count?"),
     ("Growth Area","A list of recommended counselors. Seperate line in the ministry budget for counseling that we can really only spend on that.","",None),
     ("Strength","I don\'t know. I have a good counselor back in the states that I know I can set stuff up with but I am also a believer that mentorship and community should handle 90% of what we go to counseling for.","",None),
   ],
   "open_q":"What would make counseling more accessible or effective for you?",
   "lead_question":"Staff awareness of the counseling pathway and what barriers may be limiting access.",
  },

  {"name":"Ministry Partner Development","avg":3.48,"status":"Watch",
   "groups":[("Ministry Partner Development",[
     ("MPD","Financial pressure sometimes distracts me from focusing on ministry.","Financial pressure sometimes distracts me from focusing on ministry. [Burden — inverted]",3.79,True,[0,5,2,4,8],19),
     ("MPD","I often feel alone in carrying the responsibility of MPD.","I often feel alone in carrying the responsibility of MPD. [Burden — inverted]",3.05,True,[3,0,10,5,1],19),
     ("MPD","I know who to turn to for encouragement or accountability in my MPD journey.","I know who to turn to for encouragement or accountability in my MPD journey.",3.47,False,[0,5,3,8,3],19),
     ("MPD","I have the practical MPD tools and guidance I need to raise and maintain support for long-term ministry.","I have the practical MPD tools and guidance I need to raise and maintain support for long-term ministry.",3.68,False,[0,2,5,9,3],19),
     ("MPD","I receive valid and regular financial reports about my support team, and I routinely track changes to my finances.","I receive valid and regular financial reports about my support team, and I routinely track changes to my finances.",3.58,False,[1,2,3,11,2],19),
     ("MPD","I feel supported by my uplink or ministry team in building and maintaining my support team.","I feel supported by my uplink or ministry team in building and maintaining my support team.",3.63,False,[1,4,3,4,7],19),
     ("MPD","I am confident when sharing my ministry vision and financial needs with potential supporters.","I am confident when sharing my ministry vision and financial needs with potential supporters.",3.74,False,[0,3,4,7,5],19),
     ("MPD","I am effective when sharing my ministry vision and financial needs with potential supporters.","I am effective when sharing my ministry vision and financial needs with potential supporters.",3.95,False,[0,1,6,5,7],19),
     ("MPD","I regularly communicate with my partners to let them know how their giving and praying is making an impact.","I regularly communicate with my partners to let them know how their giving and praying is making an impact.",4.11,False,[0,1,2,9,6],18),
   ])],
   "strengths":[
     'Staff communicate ministry vision with confidence and effectiveness.',
     'Partner communication rhythms are healthy and consistent across the team.',
     'MPD tools and guidance are accessible and used.',
     'Staff report feeling effective when sharing vision and financial needs with supporters.',
   ],
   "growth":[
     'Financial pressure is a significant distraction from ministry focus and would benefit from more direct support.',
     'The sense of carrying MPD alone could be reduced through more intentional team support.',
     'Accountability and encouragement rhythms in the MPD journey vary across contexts.',
     'Stage-appropriate coaching for long-tenured staff could be more structured and consistent.',
   ],
   "leadership":["Which staff on your team are experiencing financial pressure right now, and what is your response?","How are you actively sharing the MPD load so staff do not feel they are carrying it alone?","How consistent and stage-appropriate is your MPD coaching — especially for experienced staff?","What best practices does your team share around partner communication and support building?"],
      "quotes":[
     ("Growth Area","A proper coach for my level. I\'ve been on the field for 12 years. My salary is the same as when I started. Inflation, ministry reorganization... No one has followed up or asked how MPD is going.","",None),
     ("Growth Area","Pomoc w formowaniu komunikatu o mojej roli po zmianach w Fali - jak to przedstawiać potencjalnym darczyńcom tak, żeby zachęcić ich do wsparcia kiedy rola jest jeszcze niestabilna i owoce niewidoczne.","","Help forming the message about my role after the changes in Fala — how to present it to potential donors to encourage their support when the role is still unstable and the fruit is not yet visible."),
     ("Growth Area","Best practices for ways to communicate with ministry partners (newsletters, apps, etc.). Surely with JV\'s knowledge history, I shouldn\'t have to figure all this out on my own — but that is how it has been.","",None),
     ("Growth Area","Indywidualne konsultacje raz na rok / pół roku ze specjalistą od MPD w Fali. Możliwość zadania pytań, uzyskania sugestii na temat najlepszych praktyk.","","Individual consultations once or twice a year with an MPD specialist in Fala. The ability to ask questions and get suggestions on best practices."),
     ("Growth Area","Solidny ambasador w innym stanie, zborze lub kontekście.","","A solid ambassador in another state, congregation, or context."),
     ("Growth Area","A new partner church, and new opportunities to invite churches to (other than camps).","",None),
     ("Strength","Będąc na grancie miałam świetną okazję do tego żeby bardzo pożądnie w tym wzrastać, za co jestem bardzo wdzięczna.","","Being on a grant gave me a great opportunity to grow significantly in this area, for which I am very grateful."),
   ],
   "open_q":"What is one thing that would strengthen your MPD journey right now?",
   "lead_question":"Financial pressure in the MPD journey and whether support is matched to each staff member's stage.",
  },

  {"name":"Learning & Development","avg":3.51,"status":"Healthy",
   "groups":[("Learning & Development",[
     ("L&D","I often struggle to apply Christ's strategy to daily ministry.","I often struggle to apply Christ's strategy to daily ministry. [Burden — inverted]",2.38,True,[2,11,6,2,0],21),
     ("L&D","I frequently feel unsure about how to move forward in my development.","I frequently feel unsure about how to move forward in my development. [Burden — inverted]",2.95,True,[0,9,6,4,2],21),
     ("L&D","The equipping resources available enable me to be developed in my role.","The equipping resources available enable me to be developed in my role.",3.24,False,[1,3,9,6,2],21),
     ("L&D","I am experiencing professional growth in this season.","I am experiencing professional growth in this season.",3.24,False,[1,6,3,9,2],21),
     ("L&D","I receive helpful feedback and encouragement that supports my learning and development.","I receive helpful feedback and encouragement that supports my learning and development.",3.38,False,[3,2,4,8,4],21),
     ("L&D","I am growing in healthy rhythms that help me serve others from a place of wholeness.","I am growing in healthy rhythms that help me serve others from a place of wholeness.",3.38,False,[1,2,8,8,2],21),
     ("L&D","My uplink rhythms (meetings, guidance, support) help me thrive in ministry.","My uplink rhythms (meetings, guidance, support) help me thrive in ministry.",3.71,False,[2,2,3,7,7],21),
     ("L&D","I am continually learning how Christ's strategy shapes how I lead, train, and disciple others in ministry.","I am continually learning how Christ's strategy shapes how I lead, train, and disciple others in ministry.",3.86,False,[0,3,3,9,6],21),
     ("L&D","I am experiencing personal growth in this season.","I am experiencing personal growth in this season.",4.14,False,[0,1,0,15,5],21),
   ])],
   "strengths":[
     'Staff report meaningful personal growth in this season of ministry.',
     'Christ\'s strategy is increasingly shaping how staff lead and disciple others.',
     'Uplink rhythms are experienced as supportive and helpful for growth.',
   ],
   "growth":[
     'Equipping resources could be more consistently aligned with individual roles and stages.',
     'Clarity around professional development pathways could be strengthened.',
     'Feedback and encouragement in development conversations could be more specific and regular.',
     'Healthy sustainability rhythms are present but would benefit from more intentional support.',
     'More structured guidance around next steps in development would reduce uncertainty.',
   ],
   "leadership":["How are you helping staff identify their specific next steps in professional development?","What rhythms do you use to help staff connect Christ's strategy to their daily ministry decisions?","How specific and useful is the feedback you give in development conversations?","Which staff on your team feel most stuck in their development, and what is holding them back?"],
      "quotes":[
     ("Growth Area","Szkolenie z planowania prowadzenia lidera w skali roku, jak to rozplanować w czasie oraz dobrać właściwe materiały, jak dopasować do potrzeb i stylu nauki konkretnego lidera.","","Training on annual leader development planning — how to plan it over time, choose the right materials, and match them to the needs and learning style of a specific leader."),
     ("Growth Area","How to be a good uplink. How to manage a team well.","",None),
     ("Growth Area","Marzą mi się regularne grupowe case study kilku różnych sytuacji, np. z naszego lokalnego kontekstu / kościoła / uczniów. Takie sesje mogłyby budować wspólnotę, poszerzać perspektywy i dawać mądrość wielu doradców.","","I dream of regular group case studies of different situations from our local context — church, disciples, etc. Such sessions could build community, broaden perspectives, and give the wisdom of many advisors."),
     ("Growth Area","Trudno mi powiedzieć. Mam wrażenie, że po prostu praktyczne przygotowywanie się na spotkania i moje większe spędzanie czasu z Bogiem to jest to, co mnie najbardziej rozwija.","","It\'s hard to say. I feel that simply practical preparation for meetings and spending more time with God is what develops me most."),
     ("Strength","I have lots of opportunities within and outside JV, so there is nothing else I can think of.","",None),
   ],
   "open_q":"What type of training or development would be most helpful to you in the coming year?",
   "lead_question":"Staff clarity on development direction and whether equipping resources are aligned to their roles.",
  },

  {"name":"Marriages","avg":3.55,"status":"Healthy",
   "groups":[("Marriages",[
     ("Marriages","I feel supported and encouraged by JV and my team culture to prioritize time and investment in my marriage.","I feel supported and encouraged by JV and my team culture to prioritize time and investment in my marriage.",3.09,False,[1,2,4,3,1],11),
     ("Marriages","I often feel that ministry demands leave me too depleted to invest in my marriage.","I often feel that ministry demands leave me too depleted to invest in my marriage. [Burden — inverted]",2.91,True,[1,4,2,3,1],11),
     ("Marriages","I know where to go for help if our marriage faces challenges.","I know where to go for help if our marriage faces challenges.",3.64,False,[0,1,4,4,2],11),
     ("Marriages","We are learning to navigate ministry pressures together in healthy ways.","We are learning to navigate ministry pressures together in healthy ways.",3.73,False,[0,2,2,4,3],11),
     ("Marriages","I have couples or mentors I can turn to for support.","I have couples or mentors I can turn to for support.",3.82,False,[1,1,1,4,4],11),
     ("Marriages","My team culture values and respects the importance of nurturing marriages in ministry.","My team culture values and respects the importance of nurturing marriages in ministry.",3.91,False,[1,1,1,3,5],11),
   ])],
   "strengths":[
     'Team culture affirms the importance of healthy marriages in ministry.',
     'Couples report knowing where to turn for support when challenges arise.',
     'Mentors and peer couples are accessible and experienced as helpful.',
     'Couples are growing in navigating ministry pressures together.',
   ],
   "growth":[
     'Proactive support for prioritizing marriage investment could be more consistent.',
     'Ministry depletion patterns suggest that sustainability structures for couples could be strengthened.',
     'Awareness of formal marriage support resources could be more visible across the team.',
     'More intentional encouragement to invest in marriages before pressure builds would be beneficial.',
   ],
   "leadership":["How actively are you creating space for married staff to prioritize their marriages — not just in crisis?","What do you do when you notice a couple on your team showing signs of depletion?","How do staff on your team know what formal marriage support is available from JV?","What would it look like for your team culture to more consistently communicate that marriages matter?"],
      "quotes":[
     ("Growth Area","Przymusowy, regularny sabbathical - obecnie nie ma nawet zachęty o tym by myśleć o takim rozwiązaniu. Bonus finansowy na urlop/krótki wyjazd z współmałżonkiem.","","Mandatory, regular sabbatical — currently there is not even encouragement to think about such a solution. A financial bonus for vacation or a short trip with a spouse."),
     ("Growth Area","I don\'t understand the formal support that is offered in our organization for this. I know the Hashes do something, but I don\'t know what it is or if it applies to us.","",None),
     ("Growth Area","Może jakieś szkolenia online dal małżeństw...","","Maybe some online training for married couples..."),
     ("Growth Area","A marriage counseling stipend :)","",None),
   ],
   "open_q":"What would most strengthen marriages in your ministry context?",
   "lead_question":"Whether married staff feel supported to invest in their marriages alongside ministry demands.",
  },

  {"name":"JV Women","avg":4.01,"status":"Healthy",
   "groups":[("JV Women",[
     ("Women","I struggle to understand how my gifts and role fit within my ministry context and to see them being used effectively.","I struggle to understand how my gifts and role fit within my ministry context. [Burden — inverted]",2.27,True,[2,5,3,1,0],11),
     ("Women","Often I feel isolated in ministry and lack women I can easily turn to for support, prayer, and encouragement.","I often feel isolated in ministry and lack women I can easily turn to for support, prayer, and encouragement. [Burden — inverted]",2.18,True,[3,5,1,2,0],11),
     ("Women","I feel my voice is valued in team and organizational settings.","I feel my voice is valued in team and organizational settings.",4.00,False,[0,2,1,3,5],11),
     ("Women","I feel that my organization provides clear guidance on women's roles and leadership opportunities, and that women's voices are represented in leadership conversations.","I feel that my organization provides clear guidance on women's roles and leadership opportunities.",4.00,False,[0,1,2,4,4],11),
     ("Women","I often feel disconnected from my team and uninformed about its activities and decisions.","I often feel disconnected from my team and uninformed about its activities and decisions. [Burden — inverted]",1.91,True,[3,7,0,1,0],11),
     ("Women","Our organization provides opportunities for women to encourage and support one another.","Our organization provides opportunities for women to encourage and support one another.",4.09,False,[0,1,1,5,4],11),
     ("Women","JV gatherings (such as conferences or retreats) provide a safe and nourishing environment where I feel refreshed and able to share openly.","JV gatherings provide a safe and nourishing environment where I feel refreshed and able to share openly.",4.36,False,[0,0,1,5,5],11),
   ])],
   "strengths":[
     'Women report feeling connected and supported within the organization.',
     'Trusted spaces for encouragement, connection, and prayer are accessible.',
     'Women feel valued and heard in team and organizational settings.',
     'JV gatherings provide a safe and nourishing environment for women.',
     'Women report clarity around their gifts and roles within ministry.',
   ],
   "growth":[
     'Including women\'s voices before decisions are made could be more consistent.',
     'Communication patterns across teams could be more intentional and inclusive.',
   ],
   "leadership":["How are you actively affirming the gifts and calling of women on your team?","How do you ensure women are included and informed before team decisions are made — not only after?","What would it look like to increase clarity around leadership pathways for women on your team?","How do JV gatherings in your context create genuine safety and refreshment for women?"],
      "quotes":[
     ("Growth Area","Raczej mam odczucie, że moje zdanie jest mało ważne. Pomogłoby gdyby ktoś pytał o zdanie przed podejmowaniem większych decyzji lub przynajmniej włączał w ten proces.","","I rather feel that my opinion matters little. It would help if someone asked for input before making bigger decisions, or at least included us in the process."),
     ("Strength","Rozwijając kulturę wśród kobiet i pielęgnując ją","","By developing and nurturing the culture among women."),
     ("Growth Area","Myślę że w naszym kontekście nie są odczuwalne nierówności między kobietami a mężczyznami i co jest dla kogo dostępne, ale nie przeszłam jeszcze przez ciążę i urlop macierzyński, więc zastanawiam się jak kobiety w takiej sytuacji to odczuwają.","","I think in our context there are no noticeable inequalities between women and men in terms of what is available. But I have not yet gone through pregnancy and maternity leave, so I wonder how women in that situation experience it."),
   ],
   "open_q":"What kind of support or opportunities would most help women thrive in your context?",
   "lead_question":"The consistency with which women's voices are included before team decisions are made.",
  },

  {"name":"Language & Culture","avg":4.04,"status":"Healthy",
   "groups":[
     ("Second Culture Staff (n=5)",[
       ("L&C","I receive regular accountability and helpful feedback on my progress in language learning.","I receive regular accountability and helpful feedback on my progress in language learning.",2.60,False,[1,2,1,0,1],5),
       ("L&C","I regularly feel discouraged about my pace of language learning.","I regularly feel discouraged about my pace of language learning. [Burden — inverted]",3.20,True,[1,1,0,1,2],5),
       ("L&C","I often struggle to balance ministry demands with language and culture growth.","I often struggle to balance ministry demands with language and culture growth. [Burden — missing data]",3.20,True,[1,1,1,0,2],5),
       ("L&C","I feel increasingly capable in ministry because of my language and cultural skills.","I feel increasingly capable in ministry because of my language and cultural skills.",3.60,False,[0,1,0,4,0],5),
       ("L&C","I know who to turn to for help with cultural or language learning challenges.","I know who to turn to for help with language learning challenges.",3.60,False,[0,1,1,2,1],5),
       ("L&C","I clearly understand the expectations for my progress in language and culture learning.","I clearly understand the expectations for my progress in language and culture learning.",4.00,False,[0,0,1,3,1],5),
       ("L&C","My team helps me with my language and cultural adaptation needs.","My team helps me with my language and cultural adaptation needs.",4.00,False,[0,1,0,2,2],5),
       ("L&C","I am growing in my ability to live and function daily in another culture and language.","I am growing in my ability to live and function daily in another culture and language.",4.40,False,[0,0,0,3,2],5),
     ]),
     ("First Culture Staff (n=12)",[
       ("L&C","I am aware of cultural differences on my team and I intentionally seek to understand them.","I am aware of cultural differences on my team and I intentionally seek to understand them.",4.33,False,[0,0,0,8,4],12),
       ("L&C","It is important for me to grow in my English skills because I am part of such an international organization.","It is important for me to grow in my English skills because I am part of such an international organization.",4.33,False,[0,1,1,3,7],12),
       ("L&C","I am able to switch into English and still communicate effectively on a team where my language is not everyone's first language.","I am able to switch into English and still communicate effectively on a team where my language is not everyone's first language.",4.50,False,[0,0,1,4,7],12),
       ("L&C","I regularly offer grace and encouragement to teammates who experience culture shock or slow progress in language learning.","I regularly offer grace and encouragement to teammates who experience culture shock or slow progress in language learning.",4.58,False,[0,0,0,5,7],12),
     ]),
   ],
   "strengths":[
     'Second culture staff report clarity around expectations for language and culture growth.',
     'Staff are growing in their capacity to live and function in another culture.',
     'First culture staff communicate effectively in English on multilingual teams.',
     'A culture of grace and encouragement toward second culture teammates is present.',
     'First culture staff value English growth as part of serving in an international context.',
   ],
   "growth":[
     'Accountability and feedback rhythms for language learning could be more regular and structured.',
     'Encouragement around pace of language learning could be more consistent for second culture staff.',
     'Balancing ministry demands with language and culture growth could be more intentionally supported.',
     'Ministry confidence through language is growing but would benefit from more consistent affirmation.',
   ],
   "leadership":["How are you providing regular, structured accountability for language learning with second culture staff?","What expectations have you clearly communicated to second culture staff about language progress — and how do you track them?","How does your team actively support cultural adaptation, not just language learning?","How do first culture staff model patience and grace toward teammates who are still learning?"],
      "quotes":[
     ("Growth Area","It would be helpful for me and my teammates to know what is realistic for my language and cultural growth right now. I don\'t know how to set goals. A knowledgeable language learning coach would be a dream. The plan was \'figure out what works for you\' — I never knew goals or timelines.","2nd Culture",None),
     ("Growth Area","Knowledge of expectations and timeline","2nd Culture",None),
     ("Growth Area","Szczere historie amerykanów o tym, co jest dla nich trudne we współpracy z nami i jak możemy ją usprawniać. Taki feedback jest zawsze cenny.","1st Culture","Honest stories from Americans about what is difficult for them in working with us and how we can improve. That kind of feedback is always valuable."),
     ("Growth Area","Jakieś narzędzia/badania, które pomogłyby rozmawiać o różnicach kulturowych i odróżniać różnice kulturowe od osobowościowych.","1st Culture","Some tools or research that would help talk about cultural differences and distinguish cultural differences from personality differences."),
     ("Growth Area","Większa szczerość i otwartość w prośbie o pomoc ze strony osoby 2nd culture. Nie traktowanie nas, 1st culture, jako mniej uzdolnionych, mniej kompetentnych czy mniej ważnych.","1st Culture","More openness and honesty in asking for help from the second culture person. Not treating first culture staff as less gifted, less competent, or less important."),
     ("Strength","Myślę, że działamy już skutecznie.","1st Culture","I think we are already working effectively."),
   ],
   "open_q":"2nd culture: What would most help you in your language and cultural growth?  |  1st culture: What would most help you in working more effectively in a multicultural team?",
   "lead_question":"Accountability and encouragement structures for second culture staff in language learning.",
  },

]

# ── Build combined workbook ─────────────────────────────────────────────────
total=len(DEPTS)
combined=openpyxl.Workbook()
build_overview(combined.active)
for i,d in enumerate(DEPTS,1):
    ws=combined.create_sheet()
    build(d,i,total,target_ws=ws,save=False)
tabcolor={'Healthy':'C6EFCE','Watch':'FFEB9C','Concern':'FFC7CE'}
for d,sheetname in zip(DEPTS,combined.sheetnames[1:]):
    combined[sheetname].sheet_properties.tabColor=tabcolor[d['status']]
path=os.path.join(OUT,'Poland_All_Departments_Director_Review.xlsx')
combined.save(path)
print('WROTE',os.path.basename(path))
print('DONE')
